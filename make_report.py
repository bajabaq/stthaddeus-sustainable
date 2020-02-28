#!/usr/bin/python3

import openpyxl
import os
import shutil
import sys

def read_excel_city(fname, sname, city):   
    print("read the data file for the defined city")    
    #open the workbook
    wb    = openpyxl.load_workbook(fname)
    sheet = wb[sname]

    max_row = sheet.max_row
    max_col = sheet.max_column

    #get the headers (as a key for a dict)
    city_data    = {}
    keys         = []    
    maincity_col = 0
    for i in range(1,max_col+1,1):  #add 1 b/c of range
        key = sheet.cell(row=1,column=i).value
        keys.append(key)
        if key == 'maincity':
            maincity_col = i
        #endif
    #endfor

    if 'maincity' in keys:
        found_city = False
        city_row   = 0
        for i in range(1,max_row+1,1): #add 1 b/c of range
            maincity = sheet.cell(row=i,column=maincity_col).value
            if maincity == city:
                found_city = True
                city_row   = i
            #endif
        #endor
        if found_city:
            for i in range(1,max_col+1,1): #add 1 b/c of range
                val = sheet.cell(row=city_row,column=i).value
                city_data[keys[i-1]]=val
            #endfor
        else:
            print("Error - maincity (" + city + ") not found - check your spelling.")
            sys.exit()
        #endif
    else:
        print("Error - column header 'maincity' not found - check xlsx file.")
        sys.exit()
    #endif
    return(city_data)
#enddef

def read_codebook(fname,sname):
    print("reading the codebook")    
    #open the workbook
    wb    = openpyxl.load_workbook(fname)
    sheet = wb[sname]

    max_row = sheet.max_row
    max_col = sheet.max_column

    #get the headers (as a key for a dict)
    keys          = []    
    indicator_col = 0
    for i in range(1,max_col+1,1):  #add 1 b/c of range
        key = sheet.cell(row=1,column=i).value
        keys.append(key)
        if key == 'Indicator':
            indicator_col = i
        #endif
    #endfor   

    
    code_data = {}
    if 'Indicator' in keys:
        codes = []        
        for i in range(2,max_row+1,1): #start on row 2 and add 1 b/c of range
            code = sheet.cell(row=i,column=indicator_col).value

            #translate code from Codebook to match that of Results sheet
            #I already changed some that ended in RWJ
            if code == "sdg2v4_obesityRWJ":
                code = "sdg2v4_obesity"
            #endif
            if code == "sdg3v6_mentalhealthRWJ":
                code = "sdg3v6_mentalhealth"
            #endif
            if code == "sdg3v7_overdoseRWJ":
                code = "sdg3v7_overdose"
            #endif
            if code == "sdg3v8_leRWJ":
                code = "sdg3v8_lifeExpectancy"
            #endif
            if code == "sdg4v4_HSgradRWJ":
                code = "sdg4v4_HSgrad"                
            #endif
            if code == "sdg10v7_segregationRWJ":
                code = "sdg10v7_segregation"
            #endif
            
            code_data[code]={}
            for j in range(1, max_col+1,1): #loop over each column
                code_data[code].update([(keys[j-1],sheet.cell(row=i,column=j).value)])
            #endfor
        #endfor
    else:
        print("Error - column header 'Indicator' not found - check xlsx file.")
        sys.exit()
    #endif    
    return(code_data)
#enddef

#here's where the BIG math is done
def get_color_status(indicator_code_data, city_data, indicator_code):
    city_val = city_data[indicator_code]
    print(indicator_code,city_val)
    color = "red" #default
    
    sorder    = indicator_code_data['Sort Order']
    to_orange = indicator_code_data['To Orange']
    to_yellow = indicator_code_data['To Yellow']
    to_green  = indicator_code_data['To Green']

    print(city_val, sorder, to_orange, to_yellow, to_green)

    if city_val is None:
        color = "gray"
    else:    
        if "asc" in sorder:  #ascending, higher is better
            if city_val >= to_green:
                color = "green"
            elif to_green > city_val and city_val >= to_yellow:
                color = "yellow"
            elif to_yellow > city_val and city_val >= to_orange:
                color = "orange"
            #else color remains red
            #endif
        else:                #descending, lower is better
            if city_val <= to_green:
                color = "green"
            elif to_green < city_val and city_val <= to_yellow:
                color = "yellow"
            elif to_yellow < city_val and city_val <= to_orange:
                color = "orange"
            #else color is still red        
            #endif
        #endif
    #endif
    
#    print(color)

    return(color)
#enddef

def fix_val(nval):
    if nval is None:
        nval = ""
    else:
        nval = round(nval,1)
    #endif
    return(nval)
#enddef


#----------------------------------------------------------------
# main latex generator
#----------------------------------------------------------------
def make_goal_figs(rep_dir, city_data, code_data):        

    for i in range(1,18,1):
        sdgx = "sdg"+str(i)
        print("create figure for SDG: " + sdgx)

        fig_data = []
        
        sdgxv = sdgx + "v"
        for indicator_code, v in city_data.items():
            if indicator_code.startswith(sdgxv):
                print("get data for indicator:" + indicator_code)
                indicator_dict = code_data[indicator_code]
                color = get_color_status(indicator_dict,city_data,indicator_code)

                code_data[indicator_code].update([('color',color)])
                
                val   = city_data[indicator_code]      #measured value
                nval  = city_data["n_"+indicator_code] #normalized value - used to calc goal score
                val   = fix_val(val)
                nval  = fix_val(nval)
                sdg_align = indicator_dict['SDG Alignment']
                
                fig_data.append([sdg_align,indicator_code,color,indicator_dict,nval,val])
            #endif
        #endfor

        #sort fig_data by the indicator_code
        fig_data.sort(key = lambda x: x[0])

        print("making the figure for SDG"+str(i))
        #MAKE THE FIGURE HERE
        figtex = ""
        figtex = figtex + "\\begin{table}[ht]\n"
        figtex = figtex + "\\begin{tabularx}{\\textwidth}{\n"
        for j in range(0,len(fig_data)):
            figtex = figtex + "    	            X>{\\columncolor{white}[0.5\\tabcolsep]}\n"
        #endfor
        figtex = figtex + "    	            X}\n"
        figtex = figtex + "\\toprule\n"
        figtex = figtex + "\\multirow{2}{*}{\includegraphics[height=2em]{../figs/Goal"+str(i)+".pdf}} \n"
        for k in range(0,len(fig_data)):
            iname = fig_data[k][3]['Indicator name']
            figtex = figtex + " & " + iname
        #endfor
        figtex = figtex + "\\\\ \n"
        for k in range(0,len(fig_data)):
            color  = fig_data[k][2]
            nval   = fig_data[k][4]
            val    = fig_data[k][5]
            figtex = figtex + " & " + "\\cellcolor{"+color+"} " + str(val)
        #endfor                
        figtex = figtex + "\\\\ \n"

        figtex = figtex + "\\bottomrule\n"
        figtex = figtex + "\\end{tabularx}\n"
        figtex = figtex + "\\end{table}\n"

        tfname = os.path.join(rep_dir,"tab-sdg"+str(i)+".tex")
        if os.path.exists(tfname):
            os.remove(tfname)
        #endif
        with open(tfname,"w") as fh:
            fh.writelines(figtex)
        #endwith
        print("...done making figure\n")


        print("making the text here...")
        sectex = ""        
        for k in range(0,len(fig_data)):
            target     = fig_data[k][1]
            icode_data = fig_data[k][3]
            iname  = icode_data['Indicator name']
            desc   = icode_data['Description']
            uyear  = icode_data['Year']
            units  = icode_data['Units']
            geo    = icode_data['Geographic Level']
            sdg    = icode_data['SDG']
            sdga   = icode_data['SDG Alignment']
            tval   = str(round(icode_data['Target Value'],2))
            source = icode_data['Source']
            rat    = icode_data['Threshold Rationale']
            global_ind = icode_data['Global Indicator']
            us_ind = icode_data['State Indicator']
            delta  = icode_data['Changes from 2018']

            #make the comment file for this goal (to add comments on local performance)
            cfname = os.path.join(rep_dir,"comments")
            if os.path.exists(cfname):
                pass
            else:
                os.mkdir(cfname)
            #endif
            cfname = os.path.join(cfname,"sdg"+str(sdg)+"-"+target+".tex")
            if os.path.exists(cfname):
                pass
            else:
                os.mknod(cfname)
            #endif
            
            sectex = sectex + "\\subsection{SDG " + str(sdg) + " " + sdga + " " + iname + "}\n"
            sectex = sectex + "\\begin{labeling}{DESCRIPTION  }\n"

            if u"\u2265" in desc:
                print(desc)
                desc = desc.replace(u'\u2265',"$\geq$")  #>= symbol
                print(desc)
            #endif

            if u"\u00B5" in units:
                units = units.replace(u'\u00B5',"$\mu$") #mu
            #endif
            if units == "$\mu$g/m^3":
                units = "$\mu g/ m^3$"
            #endif
            if units == "%":
                units = "\%"
            #endif
            print(units)
            
            if desc[-1] in [".","!","?"]:
                pass
            else:
                desc = desc + "."
            #endif

            source = source.replace("&", "\&")
            
            sectex = sectex + "\\item [Description] " + desc + "\n"
            #sectex = sectex + "\\item [Units] " + units + "\n"
            sectex = sectex + "\\item [Geo. Level] " + geo + "\n"
            sectex = sectex + "\\item [Source] " + source + "\n"
            sectex = sectex + "\\item [Target Val]" + tval + "\n"
            sectex = sectex + "\\item [Threshold] " + rat + "\n"
            sectex = sectex + "\\item [Global Ind] " + global_ind + "\n"            
            sectex = sectex + "\\item [USA Ind] " + us_ind + "\n"            
            sectex = sectex + "\\item [Change from 2018] " + delta + "\n"            
            sectex = sectex + "\\item [Comments] \\input{./comments/sdg"+str(sdg)+"-"+ target  +"}\n"
            sectex = sectex + "\\end{labeling}\n"

        #endfor
        sectex = sectex + "\\clearpage\n"

        sfname = os.path.join(rep_dir,"sec-sdg"+str(i)+".tex")
        if os.path.exists(sfname):
            os.remove(sfname)
        #endif
        with open(sfname,"w") as fh:
            fh.writelines(sectex)
        #endwith        
        print("...done making the text\n")
        
    #endfor
    return code_data
#enddef

#----------------------------------------------------------------
def make_summary_fig(rep_dir,city_data, code_data):
    print("TODO - make the summary table")
    print("probably call an aux routine that is used in the target_figs")

    cfmt = ">{\\columncolor{white}[0.5\\tabcolsep]}"
    
    sumtex = ""
    sumtex = sumtex + "\\begin{table}[ht]\n"
    sumtex = sumtex + "\\setlength{\\tabcolsep}{2pt}\n"
    
    sumtex = sumtex + "\\centering\n"
    sumtex = sumtex + "\\caption{XX}\n"
    sumtex = sumtex + "\\label{tab:summary}\n"
    sumtex = sumtex + "\\begin{tabular}{ "+cfmt + "\n"
    for i in range(1,18):
        sumtex = sumtex + "                c"+cfmt+"\n"
    #endfor
    sumtex = sumtex + "                c}\n"
    sumtex = sumtex + "\\toprule\n"
    sumtex = sumtex + "\\includegraphics[width=\\allgoalwidth]{../figs/AllGoals.pdf} & \n"    
    for i in range(1,17):
        sumtex = sumtex + "\\includegraphics[width=\\allgoalwidth]{../figs/Goal"+str(i)+".pdf} & \n"
    #endfor
    sumtex = sumtex + "\\includegraphics[width=\\allgoalwidth]{../figs/Goal17.pdf} \\\\ \n"

    sumtex = sumtex + "\\midrule\n"

    #the goal summary
    for i in range(0,17):
        sumtex = sumtex + "\\cellcolor{red}" + str(i) + "  & \n"
    #endfor
    sumtex = sumtex + "\\cellcolor{red}17  \\\\ \n"
    sumtex = sumtex + "\\addlinespace[3ex]\n"

    #the targets
    for i in range(0,17):
        sumtex = sumtex + "\\cellcolor{red}" + str(i) + "  & \n"
    #endfor
    sumtex = sumtex + "\\cellcolor{red}17   \\\\ \n"

    #DO MORE HERE
    
    sumtex = sumtex + "\\bottomrule\n"
    sumtex = sumtex + "\\end{tabular}\n"
    sumtex = sumtex + "\\end{table}\n"

    sfname = os.path.join(rep_dir,"summary-table.tex")
    if os.path.exists(sfname):
        os.remove(sfname)
    #endif
    with open(sfname,"w") as fh:
        fh.writelines(sumtex)
    #endwith
    print("...done making summary table")
    
#enddef

#----------------------------------------------------------------
def make_summary_fig2(rep_dir, city_data, code_data):

    sdg_i = []
    for i in range(1,18,1):
        num_cols = 2        
        sdgx = "sdg"+str(i)
        print("create figure for SDG: " + sdgx)

        fig_data = []
        
        sdgxv = sdgx + "v"
        for indicator_code, v in city_data.items():
            if indicator_code.startswith(sdgxv):
                num_cols = num_cols + 1
            #endif
        #endfor
        sdg_i.append(num_cols)
    #endfor
    max_indicators = max(sdg_i)
        
    print("TODO - make the summary table")
    print("probably call an aux routine that is used in the target_figs")

    cfmt = ">{\\columncolor{white}[0.5\\tabcolsep]}"
    
    sumtex = ""
    sumtex = sumtex + "\\begin{table}[ht]\n"
    sumtex = sumtex + "\\setlength{\\tabcolsep}{2pt}\n"
    
    sumtex = sumtex + "\\centering\n"
    sumtex = sumtex + "\\caption{Summary of Goals and Indicators}\n"
    sumtex = sumtex + "\\label{tab:summary}\n"
    sumtex = sumtex + "\\begin{tabular}{ "+cfmt + "\n"
    for i in range(1,max_indicators):
        sumtex = sumtex + "                c"+cfmt+"\n"
    #endfor
    sumtex = sumtex + "                c}\n"
    sumtex = sumtex + "\\toprule\n"
    score_all = int(round(city_data["score_sdgi"],0))
    sumtex = sumtex + "\\includegraphics[width=\\allgoalwidth]{../figs/AllGoals.pdf} & " + str(score_all) + "\\\\ \n"    
    sumtex = sumtex + "\\midrule\n"
    
    for i in range(1,18):
        if i in (14,17):
            score_sdg = ""
        else:
            score_sdg = int(round(city_data["score_sdg"+str(i)],0))
        #endif
        sumtex = sumtex + "\\includegraphics[width=\\allgoalwidth]{../figs/Goal"+str(i)+".pdf} & " + str(score_sdg)
        
        sdgx     = "sdg"+str(i)
        ind_data = []
        
        sdgxv = sdgx + "v"
        for indicator_code, v in city_data.items():
            if indicator_code.startswith(sdgxv):
                indicator_dict = code_data[indicator_code]                
                sdg_align = indicator_dict['SDG Alignment']
                color = indicator_dict['color']
                nval  = city_data["n_"+indicator_code] #normalized value - used to calc goal score
                nval  = fix_val(nval)
                ind_data.append([sdg_align,color,nval])
            #endif
        #endfor

        #sort fig_data by the indicator_code
        ind_data.sort(key = lambda x: x[0])
        for x in ind_data:
            sumtex = sumtex + "& \\cellcolor{" + x[1] + "} " + str(x[2])
        #endfor
        
        sumtex = sumtex + " \\\\ \n"
        sumtex = sumtex + "\\addlinespace[1ex]\n"
        

    #endfor

    #DO MORE HERE
    
    sumtex = sumtex + "\\bottomrule\n"
    sumtex = sumtex + "\\end{tabular}\n"
    sumtex = sumtex + "\\end{table}\n"

    sfname = os.path.join(rep_dir,"summary-table.tex")
    if os.path.exists(sfname):
        os.remove(sfname)
    #endif
    with open(sfname,"w") as fh:
        fh.writelines(sumtex)
    #endwith
    print("...done making summary table")
    
#enddef


#----------------------------------------------------------------
def get_code_data(city_data):
    data_dir  = 'data'
    fname     = '2019USCitiesIndexResults.xlsx'
    fname     = os.path.join(os.getcwd(),os.path.join(data_dir,fname))

    sname     = 'Codebook'
    
    code_data = read_codebook(fname,sname)
    
    return code_data
#enddef

#----------------------------------------------------------------
def get_city_data(city):
    data_dir  = 'data'
    fname     = '2019USCitiesIndexResults.xlsx'
    fname     = os.path.join(os.getcwd(),os.path.join(data_dir,fname))

    sname     = 'Results'
    city_data = read_excel_city(fname, sname, city)
    return city_data
#enddef

#----------------------------------------------------------------
def main(city):
    #create report diretory, if needed
    dname   = city[0:3].lower()
    rep_dir = 'report-'+dname
    rep_dir = os.path.join(os.getcwd(),rep_dir)
    print(rep_dir)
    if os.path.exists(rep_dir):
        pass
    else:
        os.mkdir(rep_dir)
    #endif

    #if it doesn't exist, copy the template report
    rep_file = os.path.join(rep_dir,dname+"-sdg.tex")
    if os.path.exists(rep_file):
        pass
    else:
        rep_temp = os.path.join('report-template','template-sdg.tex')
        shutil.copyfile(rep_temp,rep_file)
    #endif

    #get the data and generate the report
    city_data = get_city_data(city)
    code_data = get_code_data(city_data)
    
    ucode_data = make_goal_figs(rep_dir, city_data, code_data)
    #make_target_figs or words that go in each subsection
#    make_summary_fig(rep_dir,city_data,ucode_data)
    make_summary_fig2(rep_dir,city_data,ucode_data)
    
#enddef

if __name__ == "__main__":
    city = 'Augusta'
    main(city)
#endif
